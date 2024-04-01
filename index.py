import openpyxl
from openpyxl.utils import get_column_letter

def convert_text_to_hyperlink(input_file, output_file, column_number):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, min_col=column_number, max_col=column_number, values_only=True), start=2):
        cell_value = row[0]
        if cell_value:
            hyperlink = f'=HYPERLINK("{cell_value}", "{cell_value}")'
            col_letter = get_column_letter(column_number)
            ws[f'{col_letter}{row_num}'].value = hyperlink
    
    wb.save(output_file)
    wb.close()

# Example usage
input_excel = 'input.xlsx'  # Replace with your input file path
output_excel = 'output.xlsx'  # Replace with desired output file path
column_number = 2  # Column containing the links (e.g., column B)

convert_text_to_hyperlink(input_excel, output_excel, column_number)
